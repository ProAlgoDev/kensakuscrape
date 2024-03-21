import os
import time
from datetime import  datetime 
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
import urllib

from openpyxl import load_workbook, Workbook
import re
from libs.start_web_driver import start_driver
from libs.check_last_information import check_last_lawyer_information, update_last_lawyer_information
from CONSTANT import SHORT_WAIT_TIME, MID_WAIT_TIME, LONG_WAIT_TIME, KENSAKU
import numpy as np

data_result_directory = os.path.join('result_data')
os.makedirs(data_result_directory, exist_ok=True)


class Kensaku():

    def __init__(self):

        super().__init__()
        self.site_link = 'https://kensaku.shiho-shoshi.or.jp/search/member.php?pageID=1'
        self.new_data = []
        self.last_info = ''
        self.last_number = 0

    def _start_driver(self):
        driver = start_driver()

        return driver

    
    def get_information(self):
        self.driver = self._start_driver()
        self.driver.get(self.site_link)
        if os.path.exists('updatedDate.npy'):
            oldDate = np.load('updatedDate.npy')
            renewDate = self.getUpdatedDate()
            if oldDate == renewDate:
                return
        result = check_last_lawyer_information()
        newData =[] if result == False else result
        self.resultName = self.name_gen()
        if os.path.exists(self.resultName):
            self.checkResult = True
            self.oldwb = load_workbook(self.resultName)
            self.oldws = self.oldwb.worksheets[0]
        else:
            self.checkResult = False
            
        try:
            WebDriverWait(self.driver, 60).until(
                EC.visibility_of_element_located((By.ID, 'kojin'))
            )
        except Exception as e:
            self.driver.quit()
            print('access error...')

        page = 0
        while True:

            table = self.driver.find_element(By.ID, 'kojin')
            table_trs = table.find_elements(By.TAG_NAME, 'tr') 
            page += 1

            self.last_number = 0

            for idx, tr_item in enumerate(table_trs):
                if idx == 0 or idx == 1:
                    continue
                

                current_item_information = tr_item.find_element(By.CSS_SELECTOR, 'td.font_m10').text

                lines = current_item_information.split('\n')
                postal_code_pattern = r'〒(\d{3}-\d{4})'
                postal_code = re.search(postal_code_pattern, lines[0]).group(1)
                address = lines[1]
                phone = lines[2].split('：')[1].strip()
                
                office_name = tr_item.find_elements(By.TAG_NAME, 'td')[5].text
                
                

                if (idx == 2) and (page == 1):
                    self.last_info = phone
                keyPoint = False
                if result == False:
                    print("initial")
                else:
                    for i in newData:
                        if postal_code in i:
                            if phone == i[postal_code]:
                                print(phone)
                                self.last_number = idx
                                keyPoint = True
                                break
                    if keyPoint == True:
                        continue
                tempDate = {}
                tempDate[postal_code] = phone
                newData.append(tempDate)
                self.get_new_information(office_name=office_name, postal_code=postal_code, address=address, phone=phone)
                self.last_number = idx

            if self.checkResult:
                self.save_old_data()
            else:
                self.save_new_data()
            update_last_lawyer_information(site_name=KENSAKU, last_info=newData)
            if self.last_number  == len(table_trs) - 1 :
                # pagination  = driver.find_element(By.CSS_SELECTOR, 'div.pagination.pagebottom')
                # next_tag = pagination.find_element(By.XPATH, "//a[contains(@title, 'next page')]")
                # next_tag.click()
                currenturl = self.driver.current_url
                pageId = int(currenturl.split('=')[1])
                pageId +=1
                print(pageId)
                self.driver.get(f"https://kensaku.shiho-shoshi.or.jp/search/member.php?pageID={pageId}")
                
            else:
                break

        if len(newData) != 0:
            update_last_lawyer_information(site_name=KENSAKU, last_info=newData)
        date = self.getUpdatedDate()
        self.saveDate(date)
        self.driver.quit()
        time.sleep(5*60)


    def get_new_information(self, office_name, postal_code,  address, phone):
        print( office_name, postal_code,  address, phone)
        fax = ''
        email = ''
        register_number = ''
        self.new_data.append(['司法書士', office_name, postal_code,  address, phone, fax, email, register_number])
        

    def save_old_data(self):

        if len(self.new_data) == 0:
            return
        
        for row in self.new_data:
            self.oldws.append(row)
        self.oldwb.save(self.resultName)
        self.new_data = []
    def save_new_data(self):
        if len(self.new_data) == 0:
            return
        wb = Workbook()
        sheet = wb.worksheets[0]
        sheet.append(["業種", "事務所名", "郵便番号", "住所", "電話番号", "FAX番号", "メールアドレス", "登録番号"])

        for row in self.new_data:
            sheet.append(row)
        wb.save(self.resultName)
        self.new_data = []
        self.checkResult = True
        self.oldwb = load_workbook(self.resultName)
        self.oldws = self.oldwb.worksheets[0]
    def name_gen(self):
        dt_now = datetime.now()
        now_time=dt_now.strftime('%Y%m%d')
        result_data_file_name= KENSAKU +"_result_data_" + now_time + ".xlsx"                        
        result_data_file = os.path.join(data_result_directory, result_data_file_name)
        return result_data_file
    
    def getUpdatedDate(self):
        renew = WebDriverWait(self.driver,20).until(EC.presence_of_element_located((By.XPATH, "//p[@id='renew']")))
        parten = r'\d+'
        temp = ''
        temp = renew.text
        date= ''
        match = re.findall(parten,temp)
        if match:
            date = int(match[0]+match[1]+match[2])
        return date
    def saveDate(self, date):
        np.save('updatedDate.npy',date)